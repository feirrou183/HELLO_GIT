import easygui
import openpyxl as xls 

file = easygui.fileopenbox('')
excel = xls.load_workbook(file,data_only = True)
#print(excel.sheet_names())
    
sheet = excel.active
#得到第一个表单对象


#下面打印出第2列全部内容
col_range = sheet.max_column
row_range = sheet.max_row


'''
for row in sheet.iter_rows():
    for cell in row:
        print(cell.value)
'''
#print('行：',sheet.max_row,'列：',sheet.max_column

#下面将数据导入数组中
cool_excel = [None] * row_range                 #先初始化行
for i in range(row_range):
    cool_excel[i] = [0]*col_range               #初始化完成
    
for i in range(row_range):
    for j in range(col_range):
        cool_excel[i][j] = sheet.cell(row = i+1 , column = j+1).value
       
       
#此时已经获得了完整的excel表格 cool_excel[][]获得了完整的表格。下面进行数据操作


'''
实现功能的方法： 创建一个新列表cool_excel_1[][] 将处理完的数据放入
'''
#cool_excel_1初始化
cool_excel_1 = [None] * row_range                 #先初始化行
for i in range(row_range):
    cool_excel_1[i] = [0]*col_range               #初始化完成

#下面代入基本值,第一行和第二行

for i in range(col_range):
    cool_excel_1[0][i] = cool_excel[0][i]    
    cool_excel_1[1][i] = cool_excel[1][i]

#下面代入客户信息，跟单员，备注，等级，记1，记2
for i in range(row_range):
    cool_excel_1[i][0] = cool_excel[i][0]
    for j in range(1,8):                                        ####这里7指AH开始到备注的所有列数 
        cool_excel_1[i][32+j] = cool_excel[i][32+j]             ####注意这里的34是所要的累积和
'''
到这里已经完成cool_excel_1的全部初始化工作（客户，备注等信息导入）。下面进行数据操作
''' 
'''
************************************************************************************
'''
mid_data = 0.00        #mid_data是累积计算的存放器
mid_data_1 = 0.00      #mid_data_1 是防止产生负数的保存值

N = 9                  #N为需要的月份数量

for i in range(2,row_range):        #这里开始只对客户数据进行处理，故从第三行开始
    mid_data = cool_excel[i][33]
    if mid_data == None:            #这里是对None防止无法实现功能
        mid_data = 0
        continue 
    for j in range(N):              #这里的6是六个月的数据
        cool_excel_1[i][31-2*j] = cool_excel[i][31 - 2*j]
        mid_data_1 = mid_data       #这里给mid_data_1赋值保证为前一个值
        if cool_excel_1[i][31 - 2 *j] == None:      #空格为未填数据
            pass
        elif cool_excel_1[i][31 - 2*j] != None:
            mid_data -= cool_excel_1[i][31 - 2*j]
            
            
         
         
        if  mid_data <= 5 or j == N-1:         #mid_data<= 5是给个最小值    j== N-1 是指已经满N个月
            if (mid_data > 0) and (mid_data <= 5):              
                break
            cool_excel_1[i][31 - 2*j] = mid_data_1
            break
    cool_excel_1[i][33] = 0.00
    
    
    

    for k in range(N):                              #累积和
        if cool_excel_1[i][31 - 2*k] == None:
            pass
        else:
            cool_excel_1[i][33] += cool_excel_1[i][31 - 2*k]
'''
***************************************************************************************
''' 
                   
#这里已经完成了对数据的处理操作，下面将cool_excel_1的全部数据导入一个新表中
easygui.msgbox(msg = '文件处理完成，请选择一个新的路径（推荐桌面）和填写新文件名')
file_new_path = easygui.filesavebox('请选择文件保存路径')

#下面将数组导入表中
for i in range(row_range):
    for j in range(col_range):
        sheet.cell(row = i+1 , column = j+1).value = cool_excel_1[i][j]
       
excel.save(file_new_path + '.xlsx')


